from openpyxl import load_workbook

#Acessando e manipulando as células
wb = load_workbook('exemplo.xlsx')
wb

sheetnames = wb.sheetnames
print(sheetnames)
sheet1 = wb['Sheet1'] 
print(sheet1)
#sheet1 = wb['Sheet1']["A3"].value
print(sheet1["A3"].value)

print(sheet1.cell(row=2, column=2).value)

print(sheet1.max_row)
print(sheet1.max_column)

for i in range(0, sheet1.max_row):
    print(sheet1.cell(row=i+1, column=2).value)

sheet1.cell(row=2, column=3).value = 75

wb.save('exemplo.xlsx')

#Agrupamento

#sheet1.merge_cells("A1:D1")
#sheet1.unmerge_cells("A1:D1")
#wb.save('exemplo.xlsx')

#Inserindo linha
#sheet1.insert_rows(4)
#sheet1.delete_rows(4)
#sheet1.delete_cols(2,5) #deleta da coluna B até F
#wb.save('exemplo.xlsx')

#Adição de imagem
from openpyxl.drawing.image import Image
img = Image('catlogo.png')
sheet1.add_image(img, 'A1')
wb.save('exemplo.xlsx')

